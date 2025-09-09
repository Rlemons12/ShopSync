"""
loader_parts.py
-----------------------------------
Load parts from your configured load sheet (CSV or XLSX)
into the `part` table using the MP2/SPC-style Part model.
"""

from __future__ import annotations
import os
import sys
import uuid
from typing import Dict, Optional, Tuple

import csv
from sqlalchemy import select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.exc import SQLAlchemyError

# --- Project imports ---
from app.modules.configuration.config import PARTS_LOADSHEET_XLSX
from app.modules.configuration.database_config import DatabaseConfig
from app.modules.configuration.log_config import logger
from app.modules.database.shopsync_db import Part

# Optional dependency for .xlsx
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------- Header normalization ----------------
HEADER_ALIASES = {
    "itemnum": "part_number",
    "description": "name",
    "oemmfg": "oem_mfg",
    "model": "model",
    "class flag": "class_flag",
    "ud6": "ud6",
    "type": "type",
    "notes": "notes",
    "specifications": "documentation",
    "asset nujmber": "_asset_number",  # ignored
    ".": "_ignored",
}


def norm_header(h: str) -> str:
    return HEADER_ALIASES.get((h or "").strip().lower(), (h or "").strip().lower())


def norm_val(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ---------------- File readers ----------------
def read_csv_rows(path: str):
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [norm_header(h) for h in (reader.fieldnames or [])]
        for row in reader:
            yield {norm_header(k): norm_val(v) for k, v in row.items()}


def read_xlsx_rows(path: str, sheet_name: Optional[str] = None):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl required for XLSX")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = [norm_header(c or "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {h: norm_val(v) for h, v in zip(headers, row)}


def iter_rows(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv"):
        return read_csv_rows(path)
    elif ext in (".xlsx",):
        return read_xlsx_rows(path)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")


# ---------------- Loader ----------------
def row_to_kwargs(row: Dict[str, Optional[str]]) -> Dict:
    return {
        "part_number": row.get("part_number"),
        "name": row.get("name"),
        "oem_mfg": row.get("oem_mfg"),
        "model": row.get("model"),
        "class_flag": row.get("class_flag"),
        "ud6": row.get("ud6"),
        "type": row.get("type"),
        "notes": row.get("notes"),
        "documentation": row.get("documentation"),
    }


def load_parts(path: str) -> Dict[str, int]:
    db = DatabaseConfig()
    engine = db.get_engine()
    SessionLocal = db.get_main_sessionmaker()
    req_id = f"REQ-{uuid.uuid4().hex[:8]}"

    stats = dict(seen=0, created=0, skipped=0, errors=0)

    with SessionLocal() as session:
        for i, row in enumerate(iter_rows(path), start=2):
            stats["seen"] += 1
            kwargs = row_to_kwargs(row)

            if not kwargs["part_number"] or not kwargs["name"]:
                logger.warning(f"[{req_id}] Row {i} missing required fields")
                stats["skipped"] += 1
                continue

            try:
                existing = session.execute(
                    select(Part).where(Part.part_number == kwargs["part_number"])
                ).scalar_one_or_none()

                if existing:
                    stats["skipped"] += 1
                    continue

                obj = Part(**kwargs)
                session.add(obj)
                session.commit()
                stats["created"] += 1
                logger.info(f"[{req_id}] Created Part {kwargs['part_number']} - {kwargs['name']}")

            except SQLAlchemyError as e:
                session.rollback()
                stats["errors"] += 1
                logger.exception(f"[{req_id}] Row {i} failed: {e}")

    logger.info(f"[{req_id}] Load complete. Stats: {stats}")
    return stats


if __name__ == "__main__":
    stats = load_parts(PARTS_LOADSHEET_XLSX)
    print("Summary:", stats)
